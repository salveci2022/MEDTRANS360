"""
MEDTRANS 360 — Plataforma Premium de Transporte de Pacientes
Versão 1.0 PRO | SpyNet Tecnologia Forense
CNPJ: 64.000.808/0001-51
"""
from flask import (Flask, render_template, request, jsonify, redirect,
                   url_for, session, send_file)
import os, json, hashlib, secrets, logging
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from functools import wraps
from collections import defaultdict
import time

from database import init_db, get_conn, audit

app = Flask(__name__)
SECRET_KEY = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.secret_key = SECRET_KEY
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.environ.get('FLASK_ENV') == 'production',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=10),
)

TZ = ZoneInfo(os.environ.get('APP_TZ', 'America/Sao_Paulo'))
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger('medtrans')

# ── Rate limiting ─────────────────────────────────────────
_rate = defaultdict(list)
def rate_limit(key, max_calls=60, window=60):
    now = time.time()
    _rate[key] = [t for t in _rate[key] if now - t < window]
    if len(_rate[key]) >= max_calls: return True
    _rate[key].append(now); return False

def get_ip():
    return (request.headers.get('X-Forwarded-For','').split(',')[0].strip()
            or request.remote_addr or '0.0.0.0')

def sanitize(v, maxlen=300):
    if not isinstance(v, str): v = str(v) if v is not None else ''
    for c in ['<','>','"',"'",'\\','\x00']: v = v.replace(c,'')
    return v.strip()[:maxlen]

def hash_senha(s): return hashlib.sha256(s.encode()).hexdigest()

# ── Segurança headers ─────────────────────────────────────
@app.after_request
def security_headers(r):
    r.headers['X-Content-Type-Options'] = 'nosniff'
    r.headers['X-Frame-Options'] = 'SAMEORIGIN'
    r.headers['X-XSS-Protection'] = '1; mode=block'
    r.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return r

# ── Auth decorators ───────────────────────────────────────
def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec


def require_perfis(*perfis):
    """Permite acesso apenas aos perfis especificados."""
    def decorator(f):
        @wraps(f)
        def decorated(*a, **kw):
            if not session.get('user_id'):
                return redirect(url_for('login'))
            if session.get('perfil') not in perfis:
                return render_template('acesso_negado.html',
                    perfil_atual=session.get('perfil'),
                    perfis_permitidos=perfis), 403
            return f(*a, **kw)
        return decorated
    return decorator

def master_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if session.get('perfil') not in ('master',):
            return jsonify({'ok': False, 'error': 'Acesso negado'}), 403
        return f(*a, **kw)
    return dec

# ── Helpers ───────────────────────────────────────────────
def usuario_atual():
    return {
        'id': session.get('user_id'),
        'nome': session.get('nome'),
        'perfil': session.get('perfil'),
        'clinica_id': session.get('clinica_id'),
        'email': session.get('email'),
    }

def clinica_filter():
    """Retorna clinica_id para filtrar dados (None = master vê tudo)"""
    if session.get('perfil') == 'master':
        return None
    return session.get('clinica_id')

# ════════════════════════════════════════════════
#  ROTAS PÚBLICAS
# ════════════════════════════════════════════════
@app.route('/', methods=['GET','POST'])
def index():
    if session.get('user_id'):
        perfil = session.get('perfil', '')
        if perfil == 'motorista':
            return redirect(url_for('painel_motorista'))
        elif perfil == 'clinica':
            return redirect(url_for('painel_clinica'))
        else:
            perfil = u['perfil']
        if perfil == 'motorista':
            return redirect(url_for('painel_motorista'))
        elif perfil == 'clinica':
            return redirect(url_for('painel_clinica'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        ip = get_ip()
        if rate_limit(f'login:{ip}', 10, 300):
            return render_template('login.html', erro='Muitas tentativas. Aguarde.')
        email = sanitize(request.form.get('email',''))
        senha = request.form.get('senha','')
        if not email or not senha:
            return render_template('login.html', erro='Preencha email e senha.')
        with get_conn() as conn:
            u = conn.execute(
                "SELECT * FROM users WHERE email=? AND ativo=1", (email,)
            ).fetchone()
        if not u or u['senha_hash'] != hash_senha(senha):
            log.warning(f"Login falhou: {email} — IP: {ip}")
            return render_template('login.html', erro='Email ou senha incorretos.')
        session.permanent = True
        session['user_id']   = u['id']
        session['nome']      = u['nome']
        session['email']     = u['email']
        session['perfil']    = u['perfil']
        session['clinica_id']= u['clinica_id']
        audit(u['email'], 'LOGIN', '', ip)
        log.info(f"Login: {email} [{u['perfil']}]")
        perfil = u['perfil']
        if perfil == 'motorista':
            return redirect(url_for('painel_motorista'))
        elif perfil == 'clinica':
            return redirect(url_for('painel_clinica'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('login.html', erro=None)

@app.route('/logout')
def logout():
    audit(session.get('email',''), 'LOGOUT', '', get_ip())
    session.clear()
    return redirect(url_for('login'))


# ════════════════════════════════════════════════
#  PAINÉIS SEPARADOS POR PERFIL
# ════════════════════════════════════════════════
@app.route('/motorista')
@require_perfis('motorista','master','operador')
def painel_motorista():
    return render_template('painel_motorista.html', usuario=usuario_atual())

@app.route('/clinica-painel')
@require_perfis('clinica','master','operador')
def painel_clinica():
    return render_template('painel_clinica.html', usuario=usuario_atual())

# API exclusiva para motorista — vê apenas SUAS corridas
@app.route('/api/motorista/minhas-corridas')
@require_perfis('motorista','master','operador')
def api_minhas_corridas():
    uid = session.get('user_id')
    perfil = session.get('perfil')
    with get_conn() as conn:
        # Busca motorista vinculado ao usuário
        if perfil == 'motorista':
            mot = conn.execute(
                "SELECT id,status FROM motoristas WHERE nome=(SELECT nome FROM users WHERE id=?) AND ativo=1 LIMIT 1",
                (uid,)).fetchone()
            if not mot:
                return jsonify({'ok': True, 'corridas': [], 'status': 'offline',
                                'msg': 'Nenhum motorista vinculado ao seu usuário'})
            mid = mot['id']
            status = mot['status']
            corridas = [dict(r) for r in conn.execute("""
                SELECT c.*, p.nome as paciente_nome
                FROM corridas c
                LEFT JOIN pacientes p ON c.paciente_id=p.id
                WHERE c.motorista_id=? AND c.status NOT IN ('cancelada')
                ORDER BY c.id DESC LIMIT 50
            """, (mid,)).fetchall()]
        else:
            status = 'online'
            corridas = [dict(r) for r in conn.execute("""
                SELECT c.*, p.nome as paciente_nome
                FROM corridas c
                LEFT JOIN pacientes p ON c.paciente_id=p.id
                ORDER BY c.id DESC LIMIT 50
            """).fetchall()]
    return jsonify({'ok': True, 'corridas': corridas, 'status': status})

# API para motorista alterar PRÓPRIO status
@app.route('/api/motorista/status', methods=['POST'])
@require_perfis('motorista','master')
def api_meu_status():
    uid = session.get('user_id')
    status = sanitize((request.get_json() or {}).get('status','offline'))
    if status not in ('online','offline'): return jsonify({'ok': False}), 400
    with get_conn() as conn:
        conn.execute(
            "UPDATE motoristas SET status=? WHERE nome=(SELECT nome FROM users WHERE id=?) AND ativo=1",
            (status, uid))
    return jsonify({'ok': True, 'status': status})

@app.route('/demo')
def demo():
    return render_template('demo.html')

# ════════════════════════════════════════════════
#  DASHBOARD
# ════════════════════════════════════════════════
@app.route('/dashboard')
@require_perfis('master','operador')
def dashboard():
    return render_template('dashboard.html', usuario=usuario_atual())

@app.route('/api/dashboard/stats')
@login_required
def dashboard_stats():
    cid = clinica_filter()
    today = datetime.now(TZ).strftime('%Y-%m-%d')
    with get_conn() as conn:
        def q(sql, args=()):
            return conn.execute(sql, args).fetchone()[0]
        base = "WHERE clinica_id=?" if cid else ""
        args = (cid,) if cid else ()
        and_ = "AND clinica_id=?" if cid else ""

        total_corridas  = q(f"SELECT COUNT(*) FROM corridas {base}", args)
        corridas_hoje   = q(f"SELECT COUNT(*) FROM corridas WHERE date(created_at)=date('now') {and_}", args if cid else ())
        em_andamento    = q(f"SELECT COUNT(*) FROM corridas WHERE status='em_andamento' {and_}", args if cid else ())
        total_pacientes = q(f"SELECT COUNT(*) FROM pacientes WHERE ativo=1 {base}", args)
        motoristas_on   = q(f"SELECT COUNT(*) FROM motoristas WHERE status='online' {and_}", args if cid else ())
        faturamento     = q(f"SELECT COALESCE(SUM(valor),0) FROM corridas WHERE status='finalizada' {and_}", args if cid else ()) or 0
        total_km        = q(f"SELECT COALESCE(SUM(distancia_km),0) FROM corridas WHERE status='finalizada' {and_}", args if cid else ()) or 0
        despesas        = q(f"SELECT COALESCE(SUM(valor),0) FROM despesas {base}", args) or 0

        # Corridas por status
        rows = conn.execute(
            f"SELECT status, COUNT(*) as n FROM corridas {base} GROUP BY status", args
        ).fetchall()
        por_status = {r['status']: r['n'] for r in rows}

        # Últimas 5 corridas
        sql_ult = f"""
            SELECT c.id, c.status, c.data_agendada, c.origem, c.destino,
                   p.nome as paciente, m.nome as motorista, c.valor
            FROM corridas c
            LEFT JOIN pacientes p ON c.paciente_id=p.id
            LEFT JOIN motoristas m ON c.motorista_id=m.id
            {'WHERE c.clinica_id=?' if cid else ''}
            ORDER BY c.id DESC LIMIT 5
        """
        ultimas = [dict(r) for r in conn.execute(sql_ult, (cid,) if cid else ()).fetchall()]

    return jsonify({
        'ok': True,
        'stats': {
            'total_corridas': total_corridas,
            'corridas_hoje': corridas_hoje,
            'em_andamento': em_andamento,
            'total_pacientes': total_pacientes,
            'motoristas_online': motoristas_on,
            'faturamento': round(float(faturamento), 2),
            'total_km': round(float(total_km), 2),
            'despesas': round(float(despesas), 2),
            'lucro': round(float(faturamento) - float(despesas), 2),
            'por_status': por_status,
        },
        'ultimas_corridas': ultimas,
    })

# ════════════════════════════════════════════════
#  PACIENTES
# ════════════════════════════════════════════════
@app.route('/pacientes')
@require_perfis('master','operador')
def pacientes():
    return render_template('pacientes.html', usuario=usuario_atual())

@app.route('/api/pacientes', methods=['GET'])
@login_required
def api_pacientes_list():
    cid = clinica_filter()
    with get_conn() as conn:
        sql = "SELECT p.*, c.nome as clinica_nome FROM pacientes p LEFT JOIN clinicas c ON p.clinica_id=c.id WHERE p.ativo=1"
        args = ()
        if cid:
            sql += " AND p.clinica_id=?"; args = (cid,)
        sql += " ORDER BY p.nome"
        rows = [dict(r) for r in conn.execute(sql, args).fetchall()]
    return jsonify({'ok': True, 'pacientes': rows})

@app.route('/api/pacientes', methods=['POST'])
@login_required
def api_pacientes_create():
    d = request.get_json() or {}
    cid = clinica_filter() or d.get('clinica_id')
    try:
        with get_conn() as conn:
            cur = conn.execute("""
                INSERT INTO pacientes
                  (nome,cpf,telefone,endereco,cidade,estado,lat,lng,
                   clinica_id,acompanhante,observacoes,necessidades)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (sanitize(d.get('nome','')), sanitize(d.get('cpf','')),
                  sanitize(d.get('telefone','')), sanitize(d.get('endereco','')),
                  sanitize(d.get('cidade','')), sanitize(d.get('estado','')),
                  d.get('lat'), d.get('lng'), cid,
                  sanitize(d.get('acompanhante','')),
                  sanitize(d.get('observacoes',''), 500),
                  sanitize(d.get('necessidades',''), 500)))
            pid = cur.lastrowid
        audit(session.get('email'), 'PACIENTE_CRIADO', f'id={pid}', get_ip())
        return jsonify({'ok': True, 'id': pid})
    except Exception as e:
        log.exception("Erro ao criar paciente")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/pacientes/<int:pid>', methods=['PUT'])
@login_required
def api_pacientes_update(pid):
    d = request.get_json() or {}
    with get_conn() as conn:
        conn.execute("""UPDATE pacientes SET nome=?,cpf=?,telefone=?,endereco=?,
            cidade=?,estado=?,acompanhante=?,observacoes=?,necessidades=?
            WHERE id=?""",
            (sanitize(d.get('nome','')), sanitize(d.get('cpf','')),
             sanitize(d.get('telefone','')), sanitize(d.get('endereco','')),
             sanitize(d.get('cidade','')), sanitize(d.get('estado','')),
             sanitize(d.get('acompanhante','')),
             sanitize(d.get('observacoes',''),500),
             sanitize(d.get('necessidades',''),500), pid))
    audit(session.get('email'), 'PACIENTE_EDITADO', f'id={pid}', get_ip())
    return jsonify({'ok': True})

@app.route('/api/pacientes/<int:pid>', methods=['DELETE'])
@login_required
def api_pacientes_delete(pid):
    with get_conn() as conn:
        conn.execute("UPDATE pacientes SET ativo=0 WHERE id=?", (pid,))
    audit(session.get('email'), 'PACIENTE_REMOVIDO', f'id={pid}', get_ip())
    return jsonify({'ok': True})

# ════════════════════════════════════════════════
#  MOTORISTAS
# ════════════════════════════════════════════════
@app.route('/motoristas')
@require_perfis('master','operador')
def motoristas():
    return render_template('motoristas.html', usuario=usuario_atual())

@app.route('/api/motoristas', methods=['GET'])
@login_required
def api_motoristas_list():
    cid = clinica_filter()
    with get_conn() as conn:
        sql = "SELECT m.*, c.nome as clinica_nome FROM motoristas m LEFT JOIN clinicas c ON m.clinica_id=c.id WHERE m.ativo=1"
        args = ()
        if cid: sql += " AND m.clinica_id=?"; args = (cid,)
        sql += " ORDER BY m.nome"
        rows = [dict(r) for r in conn.execute(sql, args).fetchall()]
    return jsonify({'ok': True, 'motoristas': rows})

@app.route('/api/motoristas', methods=['POST'])
@login_required
def api_motoristas_create():
    d = request.get_json() or {}
    cid = clinica_filter() or d.get('clinica_id')
    try:
        with get_conn() as conn:
            cur = conn.execute("""
                INSERT INTO motoristas
                  (nome,cpf,cnh,telefone,veiculo,placa,consumo_km,clinica_id)
                VALUES (?,?,?,?,?,?,?,?)
            """, (sanitize(d.get('nome','')), sanitize(d.get('cpf','')),
                  sanitize(d.get('cnh','')), sanitize(d.get('telefone','')),
                  sanitize(d.get('veiculo','')), sanitize(d.get('placa','')),
                  float(d.get('consumo_km',10)), cid))
            mid = cur.lastrowid
        audit(session.get('email'), 'MOTORISTA_CRIADO', f'id={mid}', get_ip())
        return jsonify({'ok': True, 'id': mid})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/motoristas/<int:mid>', methods=['PUT'])
@login_required
def api_motoristas_update(mid):
    d = request.get_json() or {}
    with get_conn() as conn:
        conn.execute("""UPDATE motoristas SET nome=?,cpf=?,cnh=?,telefone=?,
            veiculo=?,placa=?,consumo_km=? WHERE id=?""",
            (sanitize(d.get('nome','')), sanitize(d.get('cpf','')),
             sanitize(d.get('cnh','')), sanitize(d.get('telefone','')),
             sanitize(d.get('veiculo','')), sanitize(d.get('placa','')),
             float(d.get('consumo_km',10)), mid))
    return jsonify({'ok': True})

@app.route('/api/motoristas/<int:mid>/status', methods=['POST'])
@login_required
def api_motorista_status(mid):
    status = sanitize((request.get_json() or {}).get('status','offline'))
    with get_conn() as conn:
        conn.execute("UPDATE motoristas SET status=? WHERE id=?", (status, mid))
    return jsonify({'ok': True})

@app.route('/api/motoristas/<int:mid>', methods=['DELETE'])
@login_required
def api_motoristas_delete(mid):
    with get_conn() as conn:
        conn.execute("UPDATE motoristas SET ativo=0 WHERE id=?", (mid,))
    return jsonify({'ok': True})

# ════════════════════════════════════════════════
#  CLÍNICAS
# ════════════════════════════════════════════════
@app.route('/clinicas')
@require_perfis('master','operador')
def clinicas():
    return render_template('clinicas.html', usuario=usuario_atual())

@app.route('/api/clinicas', methods=['GET'])
@login_required
def api_clinicas_list():
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM clinicas WHERE ativo=1 ORDER BY nome").fetchall()]
    return jsonify({'ok': True, 'clinicas': rows})

@app.route('/api/clinicas', methods=['POST'])
@login_required
def api_clinicas_create():
    d = request.get_json() or {}
    try:
        with get_conn() as conn:
            cur = conn.execute("""
                INSERT INTO clinicas
                  (nome,cnpj,responsavel,telefone,email,endereco,cidade,estado,lat,lng)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (sanitize(d.get('nome','')), sanitize(d.get('cnpj','')),
                  sanitize(d.get('responsavel','')), sanitize(d.get('telefone','')),
                  sanitize(d.get('email','')), sanitize(d.get('endereco','')),
                  sanitize(d.get('cidade','')), sanitize(d.get('estado','')),
                  d.get('lat'), d.get('lng')))
            cid = cur.lastrowid
        audit(session.get('email'), 'CLINICA_CRIADA', f'id={cid}', get_ip())
        return jsonify({'ok': True, 'id': cid})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ════════════════════════════════════════════════
#  CORRIDAS
# ════════════════════════════════════════════════
@app.route('/corridas')
@require_perfis('master','operador')
def corridas():
    return render_template('corridas.html', usuario=usuario_atual())

@app.route('/api/corridas', methods=['GET'])
@login_required
def api_corridas_list():
    cid = clinica_filter()
    status = request.args.get('status')
    with get_conn() as conn:
        sql = """SELECT c.*, p.nome as paciente_nome, m.nome as motorista_nome,
                        cl.nome as clinica_nome
                 FROM corridas c
                 LEFT JOIN pacientes p  ON c.paciente_id=p.id
                 LEFT JOIN motoristas m ON c.motorista_id=m.id
                 LEFT JOIN clinicas cl  ON c.clinica_id=cl.id
                 WHERE 1=1"""
        args = []
        if cid: sql += " AND c.clinica_id=?"; args.append(cid)
        if status: sql += " AND c.status=?"; args.append(status)
        sql += " ORDER BY c.id DESC LIMIT 200"
        rows = [dict(r) for r in conn.execute(sql, args).fetchall()]
    return jsonify({'ok': True, 'corridas': rows})

@app.route('/api/corridas', methods=['POST'])
@login_required
def api_corridas_create():
    d = request.get_json() or {}
    cid = clinica_filter() or d.get('clinica_id')
    try:
        with get_conn() as conn:
            cur = conn.execute("""
                INSERT INTO corridas
                  (paciente_id,motorista_id,clinica_id,tipo,origem,destino,
                   lat_origem,lng_origem,lat_destino,lng_destino,
                   data_agendada,hora_saida,valor,observacoes,status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'agendada')
            """, (d.get('paciente_id'), d.get('motorista_id'), cid,
                  sanitize(d.get('tipo','ida')),
                  sanitize(d.get('origem','')), sanitize(d.get('destino','')),
                  d.get('lat_origem'), d.get('lng_origem'),
                  d.get('lat_destino'), d.get('lng_destino'),
                  sanitize(d.get('data_agendada','')),
                  sanitize(d.get('hora_saida','')),
                  float(d.get('valor',0)),
                  sanitize(d.get('observacoes',''),500)))
            rid = cur.lastrowid
        audit(session.get('email'), 'CORRIDA_CRIADA', f'id={rid}', get_ip())
        return jsonify({'ok': True, 'id': rid})
    except Exception as e:
        log.exception("Erro ao criar corrida")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/corridas/<int:rid>/status', methods=['POST'])
@login_required
def api_corrida_status(rid):
    d = request.get_json() or {}
    status = sanitize(d.get('status',''))
    now = datetime.now(TZ).strftime('%d/%m/%Y %H:%M:%S')
    with get_conn() as conn:
        if status == 'em_andamento':
            conn.execute("""UPDATE corridas SET status='em_andamento',
                hora_saida=?, km_inicial=?, updated_at=? WHERE id=?""",
                (now, float(d.get('km_inicial',0)), now, rid))
        elif status == 'finalizada':
            km_f = float(d.get('km_final',0))
            conn.execute("""SELECT km_inicial FROM corridas WHERE id=?""", (rid,))
            row = conn.execute("SELECT km_inicial FROM corridas WHERE id=?", (rid,)).fetchone()
            km_i = float(row['km_inicial']) if row else 0
            dist = max(0, km_f - km_i)
            conn.execute("""UPDATE corridas SET status='finalizada',hora_chegada=?,
                km_final=?,distancia_km=?,updated_at=? WHERE id=?""",
                (now, km_f, dist, now, rid))
        elif status == 'cancelada':
            conn.execute("UPDATE corridas SET status='cancelada',updated_at=? WHERE id=?", (now, rid))
    audit(session.get('email'), f'CORRIDA_{status.upper()}', f'id={rid}', get_ip())
    return jsonify({'ok': True})

@app.route('/api/corridas/<int:rid>', methods=['DELETE'])
@login_required
def api_corrida_delete(rid):
    with get_conn() as conn:
        conn.execute("DELETE FROM corridas WHERE id=?", (rid,))
    return jsonify({'ok': True})

# ════════════════════════════════════════════════
#  FINANCEIRO
# ════════════════════════════════════════════════
@app.route('/financeiro')
@require_perfis('master','operador')
def financeiro():
    return render_template('financeiro.html', usuario=usuario_atual())

@app.route('/api/despesas', methods=['POST'])
@login_required
def api_despesas_create():
    d = request.get_json() or {}
    cid = clinica_filter() or d.get('clinica_id')
    with get_conn() as conn:
        conn.execute("""INSERT INTO despesas(corrida_id,clinica_id,tipo,descricao,valor,data)
            VALUES(?,?,?,?,?,?)""",
            (d.get('corrida_id'), cid, sanitize(d.get('tipo','')),
             sanitize(d.get('descricao','')), float(d.get('valor',0)),
             sanitize(d.get('data',''))))
    return jsonify({'ok': True})

@app.route('/api/financeiro/resumo')
@login_required
def api_financeiro_resumo():
    cid = clinica_filter()
    mes = request.args.get('mes', datetime.now(TZ).strftime('%Y-%m'))
    with get_conn() as conn:
        and_ = "AND clinica_id=?" if cid else ""
        args = (cid,) if cid else ()
        fat = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) FROM corridas WHERE status='finalizada' AND strftime('%Y-%m',created_at)=? {and_}",
            (mes,)+args).fetchone()[0] or 0
        desp = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) FROM despesas WHERE strftime('%Y-%m',created_at)=? {and_}",
            (mes,)+args).fetchone()[0] or 0
        total_corr = conn.execute(
            f"SELECT COUNT(*) FROM corridas WHERE status='finalizada' AND strftime('%Y-%m',created_at)=? {and_}",
            (mes,)+args).fetchone()[0]
        km = conn.execute(
            f"SELECT COALESCE(SUM(distancia_km),0) FROM corridas WHERE status='finalizada' AND strftime('%Y-%m',created_at)=? {and_}",
            (mes,)+args).fetchone()[0] or 0
        rows_desp = conn.execute(
            f"SELECT tipo,SUM(valor) as total FROM despesas WHERE strftime('%Y-%m',created_at)=? {and_} GROUP BY tipo",
            (mes,)+args).fetchall()
    return jsonify({
        'ok': True,
        'faturamento': round(float(fat),2),
        'despesas': round(float(desp),2),
        'lucro': round(float(fat)-float(desp),2),
        'corridas': total_corr,
        'km_total': round(float(km),2),
        'custo_km': round(float(desp)/max(float(km),1),2),
        'despesas_por_tipo': [dict(r) for r in rows_desp],
    })

# ════════════════════════════════════════════════
#  RELATÓRIO PDF
# ════════════════════════════════════════════════
@app.route('/relatorio/corrida/<int:rid>')
@login_required
def relatorio_corrida_pdf(rid):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as rl
        from reportlab.lib import colors
        from io import BytesIO

        with get_conn() as conn:
            c = conn.execute("""
                SELECT cor.*, p.nome as pac_nome, p.cpf as pac_cpf,
                       p.telefone as pac_tel, p.endereco as pac_end,
                       m.nome as mot_nome, m.cnh, m.veiculo, m.placa, m.telefone as mot_tel,
                       cl.nome as cli_nome, cl.cnpj as cli_cnpj, cl.telefone as cli_tel
                FROM corridas cor
                LEFT JOIN pacientes p  ON cor.paciente_id=p.id
                LEFT JOIN motoristas m ON cor.motorista_id=m.id
                LEFT JOIN clinicas cl  ON cor.clinica_id=cl.id
                WHERE cor.id=?
            """, (rid,)).fetchone()
        if not c:
            return "Corrida não encontrada", 404
        c = dict(c)

        buf = BytesIO()
        cv = rl.Canvas(buf, pagesize=A4)
        W, H = A4
        now = datetime.now(TZ)

        # Header
        cv.setFillColorRGB(0.02, 0.08, 0.2)
        cv.rect(0, H-80, W, 80, fill=1, stroke=0)
        cv.setFillColorRGB(0, 0.6, 1)
        cv.setFont("Helvetica-Bold", 20)
        cv.drawString(40, H-42, "MEDTRANS 360")
        cv.setFillColorRGB(0.8, 0.9, 1)
        cv.setFont("Helvetica", 10)
        cv.drawString(40, H-58, "Plataforma de Transporte de Pacientes")
        cv.setFillColorRGB(0, 0.9, 0.6)
        cv.setFont("Helvetica-Bold", 10)
        cv.drawRightString(W-40, H-42, f"Corrida #{c['id']}")
        cv.setFillColorRGB(0.7, 0.8, 0.9)
        cv.setFont("Helvetica", 9)
        cv.drawRightString(W-40, H-56, f"Gerado: {now.strftime('%d/%m/%Y %H:%M')}")
        cv.setStrokeColorRGB(0, 0.5, 0.9)
        cv.setLineWidth(2); cv.line(0, H-82, W, H-82)

        y = H - 110
        def sec_title(titulo, yy):
            cv.setFillColorRGB(0.02, 0.08, 0.2)
            cv.rect(40, yy-14, W-80, 18, fill=1, stroke=0)
            cv.setFillColorRGB(0, 0.6, 1)
            cv.setFont("Helvetica-Bold", 10)
            cv.drawString(48, yy-8, titulo)
            return yy - 30

        def campo(label, val, x, yy, w=230):
            cv.setFillColorRGB(0.5, 0.5, 0.6)
            cv.setFont("Helvetica", 8)
            cv.drawString(x, yy+2, label)
            cv.setFillColorRGB(0.1, 0.1, 0.2)
            cv.setFont("Helvetica", 9)
            cv.drawString(x, yy-10, str(val or '—'))

        # Status badge
        status_colors = {
            'finalizada': (0.1, 0.6, 0.3),
            'em_andamento': (0.8, 0.5, 0),
            'agendada': (0.1, 0.4, 0.8),
            'cancelada': (0.7, 0.1, 0.1),
        }
        sc = status_colors.get(c.get('status',''), (0.4, 0.4, 0.4))
        cv.setFillColorRGB(*sc)
        cv.roundRect(W-130, y-2, 90, 20, 4, fill=1, stroke=0)
        cv.setFillColorRGB(1,1,1); cv.setFont("Helvetica-Bold", 9)
        cv.drawCentredString(W-85, y+8, (c.get('status') or '').upper())

        # Paciente
        y = sec_title("PACIENTE", y)
        campo("Nome", c.get('pac_nome'), 45, y)
        campo("CPF", c.get('pac_cpf'), 290, y)
        campo("Telefone", c.get('pac_tel'), 430, y, 120)
        y -= 24
        campo("Endereço", c.get('pac_end'), 45, y, 500)
        y -= 30

        # Motorista
        y = sec_title("MOTORISTA", y)
        campo("Nome", c.get('mot_nome'), 45, y)
        campo("CNH", c.get('cnh'), 290, y)
        campo("Telefone", c.get('mot_tel'), 430, y, 120)
        y -= 24
        campo("Veículo", c.get('veiculo'), 45, y)
        campo("Placa", c.get('placa'), 290, y)
        y -= 30

        # Clínica
        y = sec_title("CLÍNICA / EMPRESA", y)
        campo("Nome", c.get('cli_nome'), 45, y)
        campo("CNPJ", c.get('cli_cnpj'), 290, y)
        campo("Telefone", c.get('cli_tel'), 430, y, 120)
        y -= 30

        # Transporte
        y = sec_title("DADOS DO TRANSPORTE", y)
        campo("Origem", c.get('origem'), 45, y)
        campo("Destino", c.get('destino'), 290, y)
        y -= 24
        campo("Data", c.get('data_agendada'), 45, y)
        campo("Saída", c.get('hora_saida'), 200, y)
        campo("Chegada", c.get('hora_chegada'), 310, y)
        campo("Tipo", c.get('tipo','ida').upper(), 430, y, 120)
        y -= 24
        campo("KM Inicial", c.get('km_inicial',0), 45, y)
        campo("KM Final", c.get('km_final',0), 200, y)
        campo("Distância", f"{c.get('distancia_km',0):.1f} km", 310, y)
        y -= 30

        # Financeiro
        y = sec_title("RESUMO FINANCEIRO", y)
        cv.setFillColorRGB(0.04, 0.14, 0.08)
        cv.roundRect(45, y-36, W-90, 44, 6, fill=1, stroke=0)
        cv.setFillColorRGB(0, 0.9, 0.5); cv.setFont("Helvetica-Bold", 22)
        val = c.get('valor', 0)
        cv.drawCentredString(W/2, y-18, f"R$ {float(val):.2f}")
        cv.setFillColorRGB(0.5, 0.8, 0.6); cv.setFont("Helvetica", 9)
        cv.drawCentredString(W/2, y-32, "Valor do Serviço")
        y -= 56

        if c.get('observacoes'):
            y = sec_title("OBSERVAÇÕES", y)
            cv.setFillColorRGB(0.2, 0.2, 0.3); cv.setFont("Helvetica", 9)
            cv.drawString(45, y+4, str(c.get('observacoes',''))[:120])
            y -= 30

        # Assinatura
        y = max(y, 120)
        cv.setStrokeColorRGB(0.3, 0.3, 0.4); cv.setLineWidth(0.5)
        cv.line(45, y-10, W/2-20, y-10)
        cv.line(W/2+20, y-10, W-45, y-10)
        cv.setFillColorRGB(0.5, 0.5, 0.6); cv.setFont("Helvetica", 8)
        cv.drawCentredString(W/4, y-20, "Assinatura do Motorista")
        cv.drawCentredString(3*W/4, y-20, "Assinatura do Paciente/Responsável")

        # Footer
        cv.setStrokeColorRGB(0.2, 0.2, 0.3); cv.setLineWidth(0.5)
        cv.line(40, 38, W-40, 38)
        cv.setFillColorRGB(0.4, 0.4, 0.5); cv.setFont("Helvetica", 7.5)
        cv.drawString(40, 26, "MEDTRANS 360 © SpyNet Tecnologia Forense | CNPJ 64.000.808/0001-51")
        cv.drawRightString(W-40, 26, "Documento confidencial")

        cv.save(); buf.seek(0)
        fname = f"medtrans-corrida-{rid}-{now.strftime('%Y%m%d')}.pdf"
        return send_file(buf, mimetype='application/pdf', download_name=fname)
    except Exception as e:
        log.exception("Erro ao gerar PDF")
        return f"Erro ao gerar PDF: {e}", 500

@app.route('/relatorio/mensal')
@login_required
def relatorio_mensal():
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as rl
    from io import BytesIO
    cid = clinica_filter()
    mes = request.args.get('mes', datetime.now(TZ).strftime('%Y-%m'))
    with get_conn() as conn:
        and_ = "AND c.clinica_id=?" if cid else ""
        args = (mes, cid) if cid else (mes,)
        corridas = [dict(r) for r in conn.execute(f"""
            SELECT c.*, p.nome as pac_nome, m.nome as mot_nome
            FROM corridas c
            LEFT JOIN pacientes p ON c.paciente_id=p.id
            LEFT JOIN motoristas m ON c.motorista_id=m.id
            WHERE strftime('%Y-%m', c.created_at)=? {and_}
            ORDER BY c.id DESC
        """, args).fetchall()]
        fat = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) FROM corridas WHERE status='finalizada' AND strftime('%Y-%m',created_at)=? {'AND clinica_id=?' if cid else ''}",
            (mes, cid) if cid else (mes,)).fetchone()[0] or 0
        desp = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) FROM despesas WHERE strftime('%Y-%m',created_at)=? {'AND clinica_id=?' if cid else ''}",
            (mes, cid) if cid else (mes,)).fetchone()[0] or 0

    buf = BytesIO()
    cv = rl.Canvas(buf, pagesize=A4)
    W, H = A4
    now = datetime.now(TZ)

    cv.setFillColorRGB(0.02, 0.08, 0.2)
    cv.rect(0, H-75, W, 75, fill=1, stroke=0)
    cv.setFillColorRGB(0, 0.6, 1); cv.setFont("Helvetica-Bold", 18)
    cv.drawString(40, H-38, "MEDTRANS 360 — Relatório Mensal")
    cv.setFillColorRGB(0.7, 0.8, 0.9); cv.setFont("Helvetica", 10)
    cv.drawString(40, H-55, f"Período: {mes} | Gerado: {now.strftime('%d/%m/%Y %H:%M')}")
    cv.setStrokeColorRGB(0, 0.5, 0.9); cv.setLineWidth(2); cv.line(0, H-78, W, H-78)

    y = H - 105
    # Resumo boxes
    boxes = [
        (f"R$ {float(fat):.0f}", "Faturamento", (0.04,0.2,0.08)),
        (f"R$ {float(desp):.0f}", "Despesas", (0.2,0.04,0.04)),
        (f"R$ {float(fat)-float(desp):.0f}", "Lucro", (0.04,0.1,0.25)),
        (str(len(corridas)), "Corridas", (0.1,0.04,0.2)),
    ]
    bw = (W-80-30)/4
    for i,(val,lbl,bg) in enumerate(boxes):
        bx = 40+i*(bw+10)
        cv.setFillColorRGB(*bg); cv.roundRect(bx, y-48, bw, 48, 5, fill=1, stroke=0)
        cv.setFillColorRGB(0.6,0.7,0.9); cv.setFont("Helvetica",8)
        cv.drawCentredString(bx+bw/2, y-10, lbl)
        cv.setFillColorRGB(1,1,1); cv.setFont("Helvetica-Bold",16)
        cv.drawCentredString(bx+bw/2, y-32, str(val))
    y -= 68

    # Tabela
    cv.setFillColorRGB(0.04,0.08,0.22); cv.rect(40, y-14, W-80, 16, fill=1, stroke=0)
    cv.setFillColorRGB(0,0.7,1); cv.setFont("Helvetica-Bold", 7.5)
    for txt, xp in [("#",45),("Data",62),("Paciente",120),("Motorista",240),
                    ("Origem",330),("Distância",420),("Valor",470),("Status",510)]:
        cv.drawString(xp, y-9, txt)
    y -= 16

    for i, co in enumerate(corridas[:100]):
        if y < 60:
            cv.showPage()
            cv.setFillColorRGB(0.04,0.08,0.22); cv.rect(40, H-50, W-80, 16, fill=1, stroke=0)
            y = H-70
        cv.setFillColorRGB(0.06,0.1,0.2 if i%2==0 else 0.04)
        cv.rect(40, y-12, W-80, 14, fill=1, stroke=0)
        cv.setFillColorRGB(0.83,0.87,0.95); cv.setFont("Helvetica", 7.5)
        status_c = {'finalizada':(0.2,0.8,0.4),'em_andamento':(0.9,0.6,0),'agendada':(0.3,0.6,1),'cancelada':(1,0.3,0.3)}
        sc = status_c.get(co.get('status',''), (0.6,0.6,0.6))
        data_fmt = (co.get('data_agendada') or co.get('created_at',''))[:10]
        cv.drawString(45, y-8, str(co.get('id','')))
        cv.drawString(62, y-8, data_fmt)
        cv.drawString(120, y-8, str(co.get('pac_nome',''))[:15])
        cv.drawString(240, y-8, str(co.get('mot_nome',''))[:12])
        cv.drawString(330, y-8, str(co.get('origem',''))[:10])
        cv.drawString(420, y-8, f"{float(co.get('distancia_km',0)):.1f}km")
        cv.drawString(470, y-8, f"R${float(co.get('valor',0)):.0f}")
        cv.setFillColorRGB(*sc)
        cv.drawString(510, y-8, str(co.get('status',''))[:10])
        y -= 14

    cv.save(); buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f"medtrans-mensal-{mes}.pdf")

# ════════════════════════════════════════════════
#  EXPORTAÇÃO EXCEL
# ════════════════════════════════════════════════
@app.route('/export/corridas.xlsx')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    cid = clinica_filter()
    with get_conn() as conn:
        sql = """SELECT c.id, c.data_agendada, c.status, c.tipo,
                        p.nome as paciente, m.nome as motorista, cl.nome as clinica,
                        c.origem, c.destino, c.distancia_km, c.valor,
                        c.hora_saida, c.hora_chegada, c.km_inicial, c.km_final
                 FROM corridas c
                 LEFT JOIN pacientes p  ON c.paciente_id=p.id
                 LEFT JOIN motoristas m ON c.motorista_id=m.id
                 LEFT JOIN clinicas cl  ON c.clinica_id=cl.id"""
        args = ()
        if cid: sql += " WHERE c.clinica_id=?"; args = (cid,)
        sql += " ORDER BY c.id DESC"
        rows = [dict(r) for r in conn.execute(sql, args).fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Corridas MEDTRANS 360"
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_fill = PatternFill("solid", fgColor="0A1428")
    borda = Border(left=Side(style="thin",color="CBD5E1"),right=Side(style="thin",color="CBD5E1"),
                   top=Side(style="thin",color="CBD5E1"),bottom=Side(style="thin",color="CBD5E1"))
    alt = PatternFill("solid", fgColor="EFF6FF")

    cols = ["ID","Data","Status","Tipo","Paciente","Motorista","Clínica",
            "Origem","Destino","Distância (km)","Valor (R$)","Saída","Chegada","KM Ini","KM Fim"]
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda
    ws.row_dimensions[1].height = 25

    for ri, row in enumerate(rows, 2):
        vals = [row.get('id'), row.get('data_agendada'), row.get('status'),
                row.get('tipo'), row.get('paciente'), row.get('motorista'),
                row.get('clinica'), row.get('origem'), row.get('destino'),
                row.get('distancia_km'), row.get('valor'),
                row.get('hora_saida'), row.get('hora_chegada'),
                row.get('km_inicial'), row.get('km_final')]
        fill = alt if ri % 2 == 0 else None
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = borda; c.alignment = Alignment(vertical="center")
            if fill: c.fill = fill

    for i, w in enumerate([6,12,14,8,18,18,18,18,18,14,12,12,12,10,10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    now = datetime.now(TZ)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"medtrans-corridas-{now.strftime('%Y%m%d')}.xlsx")

# ════════════════════════════════════════════════
#  HEALTH + ERROS
# ════════════════════════════════════════════════


@app.route('/api/corridas/limpar-canceladas', methods=['POST'])
@login_required
def limpar_canceladas():
    cid = clinica_filter()
    with get_conn() as conn:
        if cid:
            cur = conn.execute("DELETE FROM corridas WHERE status='cancelada' AND clinica_id=?", (cid,))
        else:
            cur = conn.execute("DELETE FROM corridas WHERE status='cancelada'")
        removidas = cur.rowcount
    audit(session.get('email'), 'CORRIDAS_CANCELADAS_REMOVIDAS', f'{removidas} registros', get_ip())
    return jsonify({'ok': True, 'removidas': removidas})

@app.route('/api/corridas/<int:rid>/duplicar', methods=['POST'])
@login_required
def duplicar_corrida(rid):
    with get_conn() as conn:
        orig = conn.execute("SELECT * FROM corridas WHERE id=?", (rid,)).fetchone()
        if not orig:
            return jsonify({'ok': False, 'error': 'Corrida não encontrada'}), 404
        orig = dict(orig)
        cur = conn.execute("""
            INSERT INTO corridas
              (paciente_id,motorista_id,clinica_id,tipo,origem,destino,
               lat_origem,lng_origem,lat_destino,lng_destino,
               data_agendada,hora_saida,valor,observacoes,status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'agendada')
        """, (orig.get('paciente_id'), orig.get('motorista_id'), orig.get('clinica_id'),
              orig.get('tipo'), orig.get('origem'), orig.get('destino'),
              orig.get('lat_origem'), orig.get('lng_origem'),
              orig.get('lat_destino'), orig.get('lng_destino'),
              orig.get('data_agendada'), orig.get('hora_saida'),
              orig.get('valor'), orig.get('observacoes')))
        new_id = cur.lastrowid
    audit(session.get('email'), 'CORRIDA_DUPLICADA', f'original={rid} nova={new_id}', get_ip())
    return jsonify({'ok': True, 'id': new_id})

@app.route('/static/sw.js')
def service_worker():
    return send_file('static/sw.js', mimetype='application/javascript')


@app.route('/reset-demo-users')
def reset_demo_users():
    """Recria os usuários demo — use apenas uma vez após deploy."""
    import hashlib
    def h(s): return hashlib.sha256(s.encode()).hexdigest()
    try:
        with get_conn() as conn:
            demos = [
                ('Administrador Master', 'admin@medtrans360.com.br',    h('Admin@2025!'),    'master'),
                ('Operador Central',     'operador@medtrans360.com.br',  h('Oper@2025!'),     'operador'),
                ('Clínica São Lucas',    'clinica@medtrans360.com.br',   h('Clinica@2025!'),  'clinica'),
                ('Motorista João',       'motorista@medtrans360.com.br', h('Motor@2025!'),    'motorista'),
            ]
            for nome, email, senha, perfil in demos:
                existing = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
                if existing:
                    conn.execute("UPDATE users SET senha_hash=?, ativo=1 WHERE email=?", (senha, email))
                else:
                    conn.execute(
                        "INSERT INTO users(nome,email,senha_hash,perfil) VALUES(?,?,?,?)",
                        (nome, email, senha, perfil))
            conn.commit()
        return """
        <html><body style="font-family:Arial;background:#060c18;color:#e8f0fe;
          display:flex;align-items:center;justify-content:center;height:100vh;margin:0">
        <div style="text-align:center;max-width:460px;padding:40px;
          background:#0d1830;border-radius:16px;border:1px solid rgba(255,255,255,.1)">
          <div style="font-size:48px;margin-bottom:16px">✅</div>
          <h2 style="color:#0066ff;font-size:1.5rem;margin-bottom:16px">Usuários reconfigurados!</h2>
          <p style="color:#94a3b8;margin-bottom:24px">Agora você pode fazer login com:</p>
          <table style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:24px">
            <tr style="background:rgba(0,102,255,.1)">
              <th style="padding:8px;text-align:left;color:#0066ff">Email</th>
              <th style="padding:8px;text-align:left;color:#0066ff">Senha</th>
              <th style="padding:8px;text-align:left;color:#0066ff">Perfil</th>
            </tr>
            <tr><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07)">admin@medtrans360.com.br</td><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07)">Admin@2025!</td><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07);color:#00c37a">Master</td></tr>
            <tr><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07)">clinica@medtrans360.com.br</td><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07)">Clinica@2025!</td><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07);color:#0066ff">Clínica</td></tr>
            <tr><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07)">motorista@medtrans360.com.br</td><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07)">Motor@2025!</td><td style="padding:8px;border-top:1px solid rgba(255,255,255,.07);color:#f59e0b">Motorista</td></tr>
          </table>
          <a href="/login" style="display:inline-block;background:linear-gradient(135deg,#0066ff,#7b42f6);
            color:#fff;text-decoration:none;padding:12px 28px;border-radius:10px;font-weight:600">
            Ir para o Login →
          </a>
        </div></body></html>
        """, 200
    except Exception as e:
        return f"<pre style='color:red'>Erro: {e}</pre>", 500

@app.route('/health')
def health():
    return jsonify({'status':'healthy','app':'MEDTRANS 360','version':'1.0-PRO',
                    'timestamp': datetime.now(TZ).isoformat()})

@app.errorhandler(404)
def e404(e):
    if request.is_json: return jsonify({'ok':False,'error':'Não encontrado'}),404
    return render_template('login.html', erro='Página não encontrada.'), 404

@app.errorhandler(500)
def e500(e):
    log.exception("Erro 500")
    return render_template('login.html', erro='Erro interno. Tente novamente.'), 500

# Inicializar banco SEMPRE — tanto gunicorn quanto direto
try:
    init_db()
    log.info("Banco MEDTRANS 360 inicializado com sucesso.")
except Exception as _e:
    log.error(f"Erro ao inicializar banco: {_e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG','false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
